from selenium import webdriver
import time
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common import keys
import openpyxl
from selenium.webdriver.common.alert import Alert
from datetime import date
import smtplib

# Getting the status from excel to start process of opening timesheet
# path = "C:\\Users\\Black Ear Studio\\PycharmProjects\\pythonProject\\sheet_details.xlsx"
path = "C:\\Users\\QA\\Pyhton_project\\Pushp_Autotest\\sheet_details.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(1,4)
# print(cell_obj.value)

status= cell_obj.value
sheet_status= "Yes"
if status== sheet_status:

   driver= webdriver.Chrome("C:/Users/QA/Pyhton_project/Pushp_Autotest/Chrome/chromedriver.exe")
   driver.maximize_window()
   driver.implicitly_wait(10)

   #xpath Declaration
   username= "//input[@id='username']"
   password= "//input[@id='password']"
   login= "//input[@id='submit']"
   time_click= "//a[@class='timesheetModule']"
   creat_sheet= "//a[contains(text(),'Create Timesheet')]"
   # time_box= "//tr[@id='new_spradid']/td[3]"

   # Login and going to desired module
   driver.get("https://bitsinglass.mydsr.co.in/home")
   driver.find_element_by_xpath(username).send_keys("pushpender")
   driver.find_element_by_xpath(password).send_keys("123456")
   # driver.get("https://testing.mydsr.co.in/home")
   # driver.find_element_by_xpath(username).send_keys("pushpindersingh@1wayit.com")
   # driver.find_element_by_xpath(password).send_keys("123456")


   driver.find_element_by_xpath(login).click()
   driver.find_element_by_xpath(time_click).click()
   driver.find_element_by_xpath(creat_sheet).click()
   #driver.find_element_by_xpath("/html/body/div[3]/div[2]/div/div/div/div/div[4]/div[3]/h4/span[2]").click()


     #Getting the current date:
   today = date.today()
   # print("Today's date:", today)
   sys_date = today.strftime("%d-%m-%Y")
   # print("d1 =", type(d1))
   print("system date is :" + sys_date)

   #Looping to check match the system date with column header date
   for i in range(1,8):
       box_click= driver.find_element_by_xpath("//tr/th["+str(i)+"]").text
       datefromportal= box_click[0:10]
       # print(datefromportal)
       if sys_date == datefromportal:
          enter_time= driver.find_element_by_xpath("//tr/td["+str(i)+"]").click()
          print(enter_time)
          # print("The value of i is :" + str(i))
          time.sleep(2)
          for i in range(1,5):
             for j in range (1,4):
                cell_obj = sheet_obj.cell(i, j)
                # print(cell_obj.value)
                if j == 1:
                   cell_obj = sheet_obj.cell(i+1, j).value
                   selct = Select(driver.find_element_by_xpath("//div[@id='timesheet-modal-form']/form/div[3]/div/div//table/tbody/tr["+str(i)+"]/td["+str(j)+"]/select"))
                   selct.select_by_visible_text(cell_obj)
                elif j == 2:
                   cell_obj = sheet_obj.cell(i+1, j).value
                   print(cell_obj)
                   driver.find_element_by_xpath("/html/body/div[7]/div/div/div[2]/div/form/div[3]/div[1]/div/div[1]/div/table/tbody/tr["+str(i)+"]/td["+str(j)+"]/textarea").send_keys(cell_obj)
                else:
                   cell_obj = sheet_obj.cell(i+1, j).value
                   print(cell_obj)
                   driver.find_element_by_xpath("/html/body/div[7]/div/div/div[2]/div/form/div[3]/div[1]/div/div[1]/div/table/tbody/tr["+str(i)+"]/td["+str(j)+"]/input").send_keys(cell_obj)

                   # if cell_obj.value == None:
                   #   break
                   #driver.quit()
   subbutton= driver.find_element_by_xpath("//input[@id='submit']").click()
   Alert(driver).accept()
   time.sleep(10)
   driver.find_element_by_xpath(creat_sheet).click()
   time.sleep(5)

   for i in range(1,8):
       #print("gyuuijhjhn")
       box_click= driver.find_element_by_xpath("//tr/th["+str(i)+"]").text
       datefromportal= box_click[0:10]
       # today = date.today()
       # sys_date = today.strftime("%d-%m-%Y")
       if sys_date == datefromportal:
          enter_time= driver.find_element_by_xpath("//tr/td["+str(i)+"]/input").get_attribute('value')
          print("The staus of the sheet is :", enter_time)
          sheet_fstatus = "Submitted"
          if enter_time == sheet_fstatus:
             recipient = "pushpender.singh@bitsinglass.com"
             message = "Congratulations !!!  Your Timesheet is submitted Successfully for Approval. "

             # msg['From'] = sender_email
             # msg['Subject'] = 'Timesheet Filling Status[' + CurrentDate1 + ']'
             # msg['To'] = receiver_email
             # msg.set_content("Hello Gagandeep,\n\n" + FilledStatus + "\n\nThanks")
             # server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
             # server.login(sender_email, Password)
             # server.send_message(msg)
             # server.quit()
             # print("Email sent successfully")


             def SendEmail(recipient, message):
                server = smtplib.SMTP("smtp.gmail.com", 587)  # 587 = port number
                server.ehlo()  # check the smtp connection
                server.starttls()  # start the conection
                server.login("pushpender1wayit@gmail.com", "pushp@1234")
                server.sendmail("pushpender1wayit@gmail.com", recipient, message)



             SendEmail(recipient, message)
             print("email sent successfully")
             driver.quit()
          else:
             print("The time sheet is not submitted")
             # driver.quit()
else:
 print ("The Status of the sheet is: 'No'")
 driver.quit()
